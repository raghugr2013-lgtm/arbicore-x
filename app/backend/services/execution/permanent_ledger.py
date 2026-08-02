"""Permanent (immutable) Institutional Cycle Ledger — READ-ONLY accounting.

Every completed cycle is FROZEN into an append-only `production_ledger`
collection the moment it reaches COMPLETE. Frozen entries are never recomputed
or overwritten — they are the permanent institutional record. Each entry has:

  • a permanent cycle id + ledger id
  • full profitability breakdown (initial capital → portal buy → BDAG acquired →
    transfer/trading/withdrawal fees → gross/net proceeds → net profit → ROI)
  • the complete lifecycle (MetaMask funding → BlockDAG purchase → BDAG received
    → exchange deposit → liquidation → USDT withdrawal → final wallet receipt)
    with a timestamp, tx hash (where available), and fund location at every stage

Exports a spreadsheet-style CSV and a true .xlsx workbook. NO fund movement.
"""
import csv
import io

from core.models import new_id, now_iso
from services import db
from services.execution import ledger as modeled_ledger
from services.execution.fees import get_fees
from services.execution.fund_tracker import FUND_LOCATION

# state → (lifecycle stage label, ledger leg holding the tx ref)
LIFECYCLE_STAGES = [
    ("CREATED", "MetaMask Funding", None),
    ("PAYMENT_SENT", "BlockDAG Purchase", "payment_tx"),
    ("BDAG_RECEIVED", "BDAG Received", "bdag_receipt"),
    ("TRANSFER_SENT", "Transfer to Exchange", "transfer_tx"),
    ("DEPOSIT_CONFIRMED", "Exchange Deposit", "exchange_deposit"),
    ("SELL_FILLED", "Order-Book Liquidation", "sell_order"),
    ("WITHDRAWAL_SUBMITTED", "USDT Withdrawal", "withdrawal"),
    ("WITHDRAWAL_CONFIRMED", "Final Wallet Receipt", "wallet_receipt"),
]

EXPORT_FIELDS = [
    "cycle_id", "frozen_at", "completed_at", "route_name", "sell_venue",
    "initial_capital_usd", "portal_buy_price", "bdag_acquired", "transfer_fee_base",
    "exchange_deposit_qty", "weighted_sell_price", "fill_levels", "gross_proceeds_usd",
    "trading_fee_usd", "withdrawal_fee_usd", "gas_fee_usd", "net_proceeds_usd",
    "net_profit_usd", "roi_pct", "fills_source",
]


def _lifecycle(cycle: dict) -> list:
    hist = cycle.get("history", [])

    def _first_ts(state):
        for h in hist:
            if h.get("state") == state:
                return h.get("ts")
        return None

    ledger_legs = cycle.get("ledger") or {}
    out = []
    for state, label, leg in LIFECYCLE_STAGES:
        ts = _first_ts(state)
        if ts is None:
            continue
        ref = (ledger_legs.get(leg) or {}).get("reference") if leg else None
        out.append({"stage": label, "state": state, "timestamp": ts,
                    "tx_hash": ref, "fund_location": FUND_LOCATION.get(state, "—")})
    return out


async def _frozen_entry(cycle: dict, fees: dict) -> dict:
    acc = await modeled_ledger._ledger_entry(cycle, fees)   # full accounting math
    return {
        "ledger_id": new_id(), "cycle_id": cycle["id"], "mode": cycle.get("mode"),
        "route_id": cycle.get("route_id"), "route_name": acc["route_name"],
        "sell_venue": acc["sell_venue"], "frozen_at": now_iso(),
        "completed_at": acc["completed_at"],
        # --- profitability breakdown (B) ---
        "initial_capital_usd": acc["investment_usd"],
        "portal_buy_price": acc["portal_buy_price"],
        "bdag_acquired": acc["bdag_acquired"],
        "transfer_fee_base": acc["transfer_fee_base"],
        "gas_fee_usd": acc["gas_fee_usd"],
        "exchange_deposit_qty": acc["exchange_deposit_qty"],
        "weighted_sell_price": acc["weighted_sell_price"],
        "fills": acc["fills"], "fill_levels": acc["fill_levels"],
        "gross_proceeds_usd": acc["gross_proceeds_usd"],
        "trading_fee_usd": acc["trading_fee_usd"],
        "withdrawal_fee_usd": acc["withdrawal_fee_usd"],
        "net_proceeds_usd": acc["wallet_received_usd"],
        "net_profit_usd": acc["net_profit_usd"],
        "roi_pct": acc["roi_pct"], "fills_source": acc["fills_source"],
        # --- full lifecycle (A) ---
        "lifecycle": _lifecycle(cycle),
        "immutable": True,
    }


async def freeze_cycle(cycle: dict) -> dict | None:
    """Freeze a single completed cycle (idempotent — never overwrites)."""
    if cycle.get("state") != "COMPLETE":
        return None
    existing = await db.production_ledger.find_one({"cycle_id": cycle["id"]}, {"_id": 0})
    if existing:
        return existing
    fees = await get_fees()
    entry = await _frozen_entry(cycle, fees)
    await db.production_ledger.insert_one(dict(entry))
    entry.pop("_id", None)
    return entry


async def backfill() -> dict:
    """Freeze every already-completed shadow cycle that is not yet in the ledger."""
    fees = await get_fees()
    cycles = await db.execution_cycles.find(
        {"mode": "shadow", "state": "COMPLETE"}, {"_id": 0}).to_list(5000)
    frozen = 0
    for c in cycles:
        if await db.production_ledger.find_one({"cycle_id": c["id"]}, {"_id": 0}):
            continue
        entry = await _frozen_entry(c, fees)
        await db.production_ledger.insert_one(dict(entry))
        frozen += 1
    return {"scanned": len(cycles), "newly_frozen": frozen}


def _row(e: dict) -> dict:
    return {k: e.get(k) for k in EXPORT_FIELDS}


async def build(limit: int = 5000) -> dict:
    entries = await db.production_ledger.find(
        {}, {"_id": 0}, sort=[("completed_at", -1)]).to_list(limit)
    # reuse the modeled aggregator on a flattened shape
    flat = [{"completed_at": e["completed_at"], "net_profit_usd": e["net_profit_usd"],
             "investment_usd": e["initial_capital_usd"]} for e in entries]
    daily, weekly, monthly = modeled_ledger._aggregate(flat)
    total_net = round(sum(e["net_profit_usd"] or 0 for e in entries), 4)
    total_inv = round(sum(e["initial_capital_usd"] or 0 for e in entries), 4)
    total_fees = round(sum((e["gas_fee_usd"] or 0) + (e["trading_fee_usd"] or 0)
                           + (e["withdrawal_fee_usd"] or 0) for e in entries), 4)
    wins = sum(1 for e in entries if (e["net_profit_usd"] or 0) > 0)
    return {
        "phase": "Permanent Institutional Ledger (immutable, append-only)",
        "summary": {
            "cycles": len(entries), "total_initial_capital_usd": total_inv,
            "total_net_profit_usd": total_net, "total_fees_usd": total_fees,
            "overall_roi_pct": round(total_net / total_inv * 100, 3) if total_inv else None,
            "avg_net_per_cycle_usd": round(total_net / len(entries), 4) if entries else None,
            "profitable_cycles": wins,
            "win_rate_pct": round(wins / len(entries) * 100, 1) if entries else None,
        },
        "entries": entries,
        "daily_pnl": daily, "weekly_pnl": weekly, "monthly_pnl": monthly,
        "note": "Immutable frozen records — one permanent entry per completed cycle, never overwritten. "
                "Sell fills modeled at freeze time. No fund movement.",
    }


async def export_csv(limit: int = 5000) -> str:
    led = await build(limit)
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
    w.writeheader()
    for e in led["entries"]:
        w.writerow(_row(e))
    buf.write("\n# DAILY PnL\nperiod,cycles,initial_capital_usd,net_profit_usd,roi_pct\n")
    for r in led["daily_pnl"]:
        buf.write(f"{r['period']},{r['cycles']},{r['investment_usd']},{r['net_profit_usd']},{r['roi_pct']}\n")
    return buf.getvalue()


async def export_xlsx(limit: int = 5000) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    led = await build(limit)
    wb = Workbook()

    # --- sheet 1: cycles ---
    ws = wb.active
    ws.title = "Cycles"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2A36")
    ws.append(EXPORT_FIELDS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for e in led["entries"]:
        r = _row(e)
        ws.append([r.get(k) for k in EXPORT_FIELDS])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    # --- sheet 2: summary ---
    ws2 = wb.create_sheet("Summary")
    s = led["summary"]
    ws2.append(["Metric", "Value"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
    for k, v in s.items():
        ws2.append([k, v])

    # --- sheet 3: daily PnL ---
    ws3 = wb.create_sheet("Daily PnL")
    ws3.append(["period", "cycles", "initial_capital_usd", "net_profit_usd", "roi_pct"])
    for c in ws3[1]:
        c.font = header_font
        c.fill = header_fill
    for r in led["daily_pnl"]:
        ws3.append([r["period"], r["cycles"], r["investment_usd"], r["net_profit_usd"], r["roi_pct"]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
